import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import io
from datetime import datetime, date
from collections import defaultdict

SHEET_NAME = "Детализация МП Инспектор"

SUBJEKT_TO_DISTRICT = {
    "УНДиПР ГУ МЧС России по Чукотскому АО": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Забайкальскому краю": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Сахалинской области": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Приморскому краю": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Еврейской АО": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Камчатскому краю": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Республике Саха (Якутия)": "Дальневосточный ФО",
    "УНДиПР Главного управления МЧС России по Республике Бурятия": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Хабаровскому краю": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Магаданской области": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Амурской области": "Дальневосточный ФО",
    "УНДиПР ГУ МЧС России по Донецкой Народной Республике": "Новые регионы",
    "УНДиПР ГУ МЧС России по Запорожской области": "Новые регионы",
    "УНДиПР ГУ МЧС России по Херсонской области": "Новые регионы",
    "УНДиПР ГУ МЧС России по Луганской Народной Республике": "Новые регионы",
    "УНДиПР ГУ МЧС России по Пензенской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Оренбургской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Ульяновской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Республике Башкортостан": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Удмуртской Республике": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Самарской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Нижегородской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Кировской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Пермскому краю": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Саратовской области": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Республике Мордовия": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Чувашской Республике - Чувашии": "Приволжский ФО",
    "УНДиПР Главного управления МЧС России по Республике Марий Эл": "Приволжский ФО",
    "УНДиПР ГУ МЧС России по Республике Татарстан": "Приволжский ФО",
    "УНДПР Главного управления МЧС России по г. Санкт-Петербургу": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Ленинградской области": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Калининградской области": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Псковской области": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Республике Коми": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Архангельской области": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Вологодской области": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Новгородской области": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Республике Карелия": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Мурманской области": "Северо-Западный ФО",
    "ОНДиПР ГУ МЧС России по Ненецкому автономному округу": "Северо-Западный ФО",
    "УНДиПР ГУ МЧС России по Кабардино-Балкарской Республике": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Республике Северная Осетия - Алания": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Республике Дагестан": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Карачаево-Черкесской Республике": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Ставропольскому краю": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Республике Ингушетия": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Чеченской Республике": "Северо-Кавказский ФО",
    "УНДиПР ГУ МЧС России по Республике Тыва": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Новосибирской области": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Кемеровской области - Кузбассу": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Красноярскому краю": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Томской области": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Республике Алтай": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Омской области": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Республике Хакасия": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Алтайскому краю": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Иркутской области": "Сибирский ФО",
    "УНДиПР ГУ МЧС России по Курганской области": "Уральский ФО",
    "УНДиПР ГУ МЧС России по Свердловской области": "Уральский ФО",
    "УНДиПР ГУ МЧС России по Челябинской области": "Уральский ФО",
    "УНДиПР ГУ МЧС России по Ямало-Ненецкому АО": "Уральский ФО",
    "УНДиПР ГУ МЧС России по Ханты-Мансийскому АО - Югре": "Уральский ФО",
    "УНДиПР ГУ МЧС России по Тюменской области": "Уральский ФО",
    "УНДиПР ГУ МЧС России по Тверской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Курской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по г. Москве": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Московской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Владимирской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Тамбовской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Тульской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Липецкой области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Рязанской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Костромской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Ярославской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Ивановской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Воронежской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Калужской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Белгородской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Брянской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по Смоленской области": "Центральный ФО",
    "УНДПР ГУ МЧС России по Орловской области": "Центральный ФО",
    "УНДиПР ГУ МЧС России по г. Севастополю": "Южный ФО",
    "УНДиПР ГУ МЧС России по Волгоградской области": "Южный ФО",
    "УНДиПР ГУ МЧС России по Ростовской области": "Южный ФО",
    "УНДиПР ГУ МЧС России по Республике Адыгея": "Южный ФО",
    "УНДиПР ГУ МЧС России по Астраханской области": "Южный ФО",
    "УНДиПР ГУ МЧС России по Республике Крым": "Южный ФО",
    "УНДиПР ГУ МЧС России по Республике Калмыкия": "Южный ФО",
    "УНДиПР ГУ МЧС России по Краснодарскому краю": "Южный ФО",
}

COLUMN_KEYWORDS = {
    'subjekt':         ['субъект рф'],
    'vid_nadzora':     ['вид надзора'],
    'nom_knm':         ['номер кнм'],
    'vid':             ['вид'],
    'status':          ['статус кнм'],
    'narusheniya':     ['нарушения выявлены'],
    'proverka_ogv':    ['проверка огв/омсу'],
    'knd':             ['кнд'],
    'ssylki':          ['ссылки на файлы'],
    'date_act':        ['дата составления акта о результате кнм', 'дата составления акта'],
    'tip_prof_vizita': ['тип проф. визита', 'тип профилактического визита'],
    's_vks':           ['с вкс', 'вкс'],
}


def normalize_str(s):
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s).strip()).lower()


def find_column_index(headers, possible_names):
    headers_norm = [normalize_str(h) for h in headers]
    possible_norm = [normalize_str(n) for n in possible_names]
    for idx, norm in enumerate(headers_norm):
        if norm in possible_norm:
            return idx
    for idx, norm in enumerate(headers_norm):
        for pname in possible_norm:
            if pname in norm:
                return idx
    return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def get_district(subj):
    if not subj:
        return "Не определено"
    return SUBJEKT_TO_DISTRICT.get(" ".join(str(subj).split()), "Не определено")


def load_data(file_bytes):
    """Принимает bytes Excel-файла, возвращает (data, col_indices, headers_orig)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Лист «{SHEET_NAME}» не найден. "
            f"Доступные листы: {', '.join(wb.sheetnames)}"
        )
    ws = wb[SHEET_NAME]
    headers = [cell.value if cell.value else "" for cell in ws[1]]
    headers_orig = headers[:]

    col_indices = {}
    missing = []
    for key, possible_names in COLUMN_KEYWORDS.items():
        idx = find_column_index(headers, possible_names)
        if idx is None:
            missing.append(f"  • «{key}» (искали: {possible_names})")
        col_indices[key] = idx

    if missing:
        raise ValueError("Не найдены обязательные столбцы:\n" + "\n".join(missing))

    data = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if not all(c is None for c in row)
    ]
    return data, col_indices, headers_orig


def filter_by_date(data, col_idx, date_from, date_to):
    date_col = col_idx['date_act']
    filtered, skipped = [], 0
    for row in data:
        parsed = parse_date(row[date_col])
        if parsed is None:
            skipped += 1
            continue
        if date_from <= parsed <= date_to:
            filtered.append(row)
        else:
            skipped += 1
    return filtered, skipped


def calculate_all_metrics(data, col_idx):
    subj_col        = col_idx['subjekt']
    knm_col         = col_idx['nom_knm']
    vid_col         = col_idx['vid']
    status_col      = col_idx['status']
    proverka_col    = col_idx['proverka_ogv']
    vid_nadzora_col = col_idx['vid_nadzora']
    knd_col         = col_idx['knd']
    nar_col         = col_idx['narusheniya']
    vks_col         = col_idx['s_vks']
    ssylki_col      = col_idx['ssylki']

    allowed_vids = {"", "выездная проверка", "рейдовый осмотр", "инспекционный визит"}
    metrics  = defaultdict(lambda: [set(), set(), set(), set(), set()])
    seen     = {k: set() for k in ('vks_denom','vks_num','och_denom','och_num','och_nar_denom','och_nar_num')}
    detail   = {k: []    for k in seen}
    denom_rows_vks, denom_rows_och = {}, {}
    rejected_vks, rejected_och    = [], []

    for row in data:
        reasons_base = []
        if normalize_str(row[status_col]) != "завершена":
            reasons_base.append(f"Статус: {row[status_col]}")
        if normalize_str(row[proverka_col]) != "нет":
            reasons_base.append(f"Проверка ОГВ: {row[proverka_col]}")
        if normalize_str(row[vid_nadzora_col]) == "гнго":
            reasons_base.append("Вид надзора: ГНГО")
        subj = row[subj_col]
        knm  = row[knm_col]
        if not subj or not knm:
            reasons_base.append("Пустой субъект или номер КНМ")

        if reasons_base:
            r = "; ".join(reasons_base)
            rejected_vks.append(tuple(row) + (r,))
            rejected_och.append(tuple(row) + (r,))
            continue

        vid_val   = normalize_str(row[vid_col])   if row[vid_col]   else ""
        knd_str   = normalize_str(row[knd_col])   if row[knd_col]   else ""
        nar_str   = normalize_str(row[nar_col])   if row[nar_col]   else ""
        vks_str   = normalize_str(row[vks_col])   if row[vks_col]   else ""
        ssylki_ok = row[ssylki_col] is not None and str(row[ssylki_col]).strip() != ""
        sk = knm

        # --- ВКС ---
        if vid_val in allowed_vids:
            metrics[subj][0].add(knm)
            if sk not in seen['vks_denom']:
                seen['vks_denom'].add(sk); detail['vks_denom'].append(row)
            if vks_str == "да" and ssylki_ok:
                metrics[subj][1].add(knm)
                if sk not in seen['vks_num']:
                    seen['vks_num'].add(sk); detail['vks_num'].append(row)
            if sk not in denom_rows_vks:
                r = []
                if vks_str != "да":   r.append(f"С ВКС ≠ 'да': {row[vks_col]}")
                if not ssylki_ok:     r.append("Ссылки пустые")
                denom_rows_vks[sk] = (row, "; ".join(r))
        else:
            rejected_vks.append(tuple(row) + (f"Вид КНМ не входит в список ВКС: {row[vid_col]}",))

        # --- Очные ---
        if vid_val in allowed_vids:
            if "осмотр" in knd_str:
                metrics[subj][2].add(knm)
                if sk not in seen['och_denom']:
                    seen['och_denom'].add(sk); detail['och_denom'].append(row)
                if vks_str == "нет" and ssylki_ok:
                    metrics[subj][3].add(knm)
                    if sk not in seen['och_num']:
                        seen['och_num'].add(sk); detail['och_num'].append(row)
                if sk not in denom_rows_och:
                    r = []
                    if vks_str != "нет": r.append(f"С ВКС ≠ 'нет': {row[vks_col]}")
                    if not ssylki_ok:    r.append("Ссылки пустые")
                    denom_rows_och[sk] = (row, "; ".join(r))
                if nar_str == "да":
                    metrics[subj][4].add(knm)
                    if sk not in seen['och_nar_denom']:
                        seen['och_nar_denom'].add(sk); detail['och_nar_denom'].append(row)
                    if vks_str == "нет" and ssylki_ok:
                        if sk not in seen['och_nar_num']:
                            seen['och_nar_num'].add(sk); detail['och_nar_num'].append(row)
            else:
                rejected_och.append(tuple(row) + (f"КНД не содержит 'осмотр': {row[knd_col]}",))
        else:
            rejected_och.append(tuple(row) + (f"Вид КНМ не входит в список очных: {row[vid_col]}",))

    denom_not_num_vks = [
        tuple(row) + (reason,)
        for knm, (row, reason) in denom_rows_vks.items()
        if knm not in seen['vks_num']
    ]
    denom_not_num_och = [
        tuple(row) + (reason,)
        for knm, (row, reason) in denom_rows_och.items()
        if knm not in seen['och_num']
    ]
    return metrics, detail, rejected_vks, rejected_och, denom_not_num_vks, denom_not_num_och


def build_report_data(metrics_dict):
    result = []
    for subj, sets in metrics_dict.items():
        result.append({
            'Субъект':       subj,
            'Округ':         get_district(subj),
            'total_vks':     len(sets[0]),
            'prim_vks':      len(sets[1]),
            'total_och':     len(sets[2]),
            'prim_och':      len(sets[3]),
            'total_och_nar': len(sets[4]),
        })
    return result


def build_excel(report_data, headers_orig, detail,
                rejected_vks, rejected_och,
                denom_not_num_vks, denom_not_num_och):
    """Формирует Excel и возвращает bytes."""
    wb = openpyxl.Workbook()

    def style_header(ws, fill_color):
        fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
        font = Font(bold=True, color="FFFFFF", name="Arial")
        for cell in ws[1]:
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 40

    def autowidth(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 60)

    def write_sheet(title, rows, color, extra_header=None):
        ws = wb.create_sheet(title=title)
        ws.append(list(headers_orig) + ([extra_header] if extra_header else []))
        style_header(ws, color)
        for row in rows:
            ws.append(list(row))
        autowidth(ws)

    # --- ВКС ---
    ws_vks = wb.active
    ws_vks.title = "ВКС"
    ws_vks.append(['№','Федеральный округ','Субъект','Всего в ААС КНД','Всего в МП Инспектор','Доля'])
    style_header(ws_vks, "4472C4")

    # --- Очные ---
    ws_och = wb.create_sheet("Очные")
    ws_och.append(['№','Федеральный округ','Субъект',
                   'Всего в ААС КНД','Всего в МП Инспектор','Доля от всех мероприятий',
                   'Всего в ААС КНД с нарушениями','Всего в МП Инспектор','Доля от мероприятий с нарушениями'])
    style_header(ws_och, "4472C4")

    districts = sorted(set(i['Округ'] for i in report_data))
    tv_all = sum(i['total_vks']    for i in report_data)
    pv_all = sum(i['prim_vks']     for i in report_data)
    to_all = sum(i['total_och']    for i in report_data)
    po_all = sum(i['prim_och']     for i in report_data)
    tn_all = sum(i['total_och_nar']for i in report_data)

    ws_vks.append(['','Итог по России','Россия', tv_all, pv_all,
                   f"{pv_all/tv_all*100:.2f}%" if tv_all else "0.00%"])
    ws_och.append(['','Итог по России','Россия',
                   to_all, po_all, f"{po_all/to_all*100:.2f}%" if to_all else "0.00%",
                   tn_all, po_all, f"{po_all/tn_all*100:.2f}%" if tn_all else "0.00%"])

    n_vks = n_och = 1
    for d in districts:
        ws_vks.append(['', d, '', '', '', ''])
        ws_och.append(['', d, '', '', '', '', '', '', ''])
        sub = [i for i in report_data if i['Округ'] == d]
        for it in sorted(sub, key=lambda x: x['prim_vks']/x['total_vks'] if x['total_vks'] else 0):
            dv = it['prim_vks']/it['total_vks']*100 if it['total_vks'] else 0
            ws_vks.append([n_vks, '', it['Субъект'], it['total_vks'], it['prim_vks'], f"{dv:.2f}%"])
            n_vks += 1
        for it in sorted(sub, key=lambda x: x['prim_och']/x['total_och_nar'] if x['total_och_nar'] else 0):
            do  = it['prim_och']/it['total_och']    *100 if it['total_och']     else 0
            don = it['prim_och']/it['total_och_nar']*100 if it['total_och_nar'] else 0
            ws_och.append([n_och, '', it['Субъект'],
                           it['total_och'], it['prim_och'], f"{do:.2f}%",
                           it['total_och_nar'], it['prim_och'], f"{don:.2f}%"])
            n_och += 1

    for ws in [ws_vks, ws_och]:
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    write_sheet("ВКС всего",          detail['vks_denom'],     "538135")
    write_sheet("ВКС с МП",           detail['vks_num'],       "C55A11")
    write_sheet("ВКС без МП",         denom_not_num_vks,       "7030A0", "Причина")
    write_sheet("ВКС — Отсеянные",    rejected_vks,            "C00000", "Причина отклонения")
    write_sheet("Очные всего",        detail['och_denom'],     "538135")
    write_sheet("Очные с МП",         detail['och_num'],       "C55A11")
    write_sheet("Очные без МП",       denom_not_num_och,       "7030A0", "Причина")
    write_sheet("Очные всего (нар.)", detail['och_nar_denom'], "375623")
    write_sheet("Очные с МП (нар.)",  detail['och_nar_num'],   "843C0C")
    write_sheet("Очные — Отсеянные",  rejected_och,            "C00000", "Причина отклонения")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def run_analysis(file_bytes, date_from, date_to):
    """
    Единая точка входа.
    file_bytes : bytes  — содержимое xlsx-файла
    date_from  : date
    date_to    : date
    Возвращает (excel_bytes, stats_dict) или бросает исключение.
    """
    data, col_idx, headers_orig = load_data(file_bytes)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    data_filtered, skipped = filter_by_date(data, col_idx, date_from, date_to)
    if not data_filtered:
        raise ValueError("За выбранный период данных нет. Проверьте даты.")

    metrics, detail, rej_vks, rej_och, dnn_vks, dnn_och = \
        calculate_all_metrics(data_filtered, col_idx)

    report_data  = build_report_data(metrics)
    excel_bytes  = build_excel(report_data, headers_orig, detail, rej_vks, rej_och, dnn_vks, dnn_och)

    tv = sum(i['total_vks']    for i in report_data)
    pv = sum(i['prim_vks']     for i in report_data)
    to = sum(i['total_och']    for i in report_data)
    po = sum(i['prim_och']     for i in report_data)
    tn = sum(i['total_och_nar']for i in report_data)

    stats = {
        'total_rows':    len(data),
        'filtered_rows': len(data_filtered),
        'skipped':       skipped,
        'subjects':      len(metrics),
        'vks_denom':     tv,
        'vks_num':       pv,
        'vks_pct':       f"{pv/tv*100:.1f}" if tv else "0.0",
        'och_denom':     to,
        'och_num':       po,
        'och_pct':       f"{po/to*100:.1f}" if to else "0.0",
        'och_nar_denom': tn,
        'och_nar_pct':   f"{po/tn*100:.1f}" if tn else "0.0",
    }
    return excel_bytes, stats
